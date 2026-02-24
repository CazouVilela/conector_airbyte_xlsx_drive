"""
Airbyte source connector for Google Sheets and XLSX files on Google Drive.

Implements the Airbyte protocol directly (spec/check/discover/read) without
the airbyte-cdk, to avoid Python version constraints.

Dual-mode: detects MIME type via Drive API and routes to either:
  - XLSX on Drive -> Download via Drive API + parse with openpyxl
  - Native Google Sheets -> Sheets API v4
"""

from __future__ import annotations

import io
import json
import logging
import re
import sys
import time as time_mod
import unicodedata
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Optional

import openpyxl
import yaml
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

logger = logging.getLogger("airbyte")

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MIME_GOOGLE_SHEETS = "application/vnd.google-apps.spreadsheet"
SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
]
SCHEMA_SAMPLE_SIZE = 1000


# ---------------------------------------------------------------------------
# Airbyte protocol message helpers
# ---------------------------------------------------------------------------

def _emit(msg: dict) -> None:
    """Write an Airbyte protocol message to stdout."""
    print(json.dumps(msg, ensure_ascii=False))


def _log(level: str, message: str) -> None:
    _emit({"type": "LOG", "log": {"level": level, "message": message}})


def _emit_record(stream: str, data: dict, emitted_at: int) -> None:
    _emit({
        "type": "RECORD",
        "record": {
            "stream": stream,
            "data": data,
            "emitted_at": emitted_at,
        },
    })


def _emit_state(data: dict) -> None:
    _emit({"type": "STATE", "state": {"data": data}})


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sanitize_column_name(name: str) -> str:
    """Convert column name to lowercase, remove accents, replace non-alnum with underscores."""
    if not name:
        return "unnamed"
    nfkd = unicodedata.normalize("NFKD", str(name))
    ascii_str = nfkd.encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^a-zA-Z0-9]", "_", ascii_str)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_").lower()
    return cleaned or "unnamed"


def deduplicate_headers(headers: list[str]) -> list[str]:
    """Append _1, _2, ... to duplicate column names."""
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


def serialize_value(val: Any) -> Any:
    """Serialize openpyxl cell values to JSON-compatible types."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, time):
        return val.isoformat()
    if isinstance(val, (int, float, bool, str)):
        return val
    return str(val)


def infer_json_type(val: Any) -> str:
    """Map a Python value to JSON Schema type."""
    if val is None:
        return "string"
    if isinstance(val, bool):
        return "boolean"
    if isinstance(val, int):
        return "integer"
    if isinstance(val, float):
        return "number"
    if isinstance(val, (datetime, date, time)):
        return "string"
    return "string"


def infer_json_format(val: Any) -> Optional[str]:
    """Return JSON Schema format hint for temporal types."""
    if isinstance(val, datetime):
        return "date-time"
    if isinstance(val, date):
        return "date"
    if isinstance(val, time):
        return "time"
    return None


def build_schema_from_sample(headers: list[str], rows: list[list[Any]]) -> dict:
    """Infer JSON schema by sampling up to SCHEMA_SAMPLE_SIZE rows."""
    properties: dict[str, dict] = {}
    sample = rows[:SCHEMA_SAMPLE_SIZE]

    for col_idx, header in enumerate(headers):
        types_seen: set[str] = set()
        format_seen: Optional[str] = None

        for row in sample:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None:
                continue
            types_seen.add(infer_json_type(val))
            fmt = infer_json_format(val)
            if fmt:
                format_seen = fmt

        if not types_seen:
            types_seen = {"string"}

        if len(types_seen) > 1:
            prop: dict[str, Any] = {"type": ["string", "null"]}
        else:
            json_type = types_seen.pop()
            prop = {"type": [json_type, "null"]}
            if format_seen:
                prop["format"] = format_seen

        properties[header] = prop

    return {
        "$schema": "http://json-schema.org/draft-07/schema#",
        "type": "object",
        "properties": properties,
    }


def strip_trailing_none_columns(
    headers: list[str], rows: list[list[Any]]
) -> tuple[list[str], list[list[Any]]]:
    """Remove trailing columns where header is empty and all data is None."""
    if not headers:
        return headers, rows

    last_valid = len(headers) - 1
    while last_valid >= 0:
        header_empty = not headers[last_valid] or headers[last_valid].strip() == ""
        if not header_empty:
            break
        all_none = all(
            (col_idx >= len(row) or row[col_idx] is None)
            for row in rows
            for col_idx in [last_valid]
        )
        if not all_none:
            break
        last_valid -= 1

    cut = last_valid + 1
    return headers[:cut], [row[:cut] for row in rows]


def filter_empty_rows(rows: list[list[Any]]) -> list[list[Any]]:
    """Remove rows where all values are None."""
    return [row for row in rows if any(v is not None for v in row)]


# ---------------------------------------------------------------------------
# Google API helpers
# ---------------------------------------------------------------------------

def _build_google_services(credentials_json: str) -> tuple[Any, Any]:
    """Build Drive and Sheets API service objects from JSON credentials."""
    creds_dict = json.loads(credentials_json)
    credentials = service_account.Credentials.from_service_account_info(
        creds_dict, scopes=SCOPES
    )
    drive_service = build("drive", "v3", credentials=credentials, cache_discovery=False)
    sheets_service = build("sheets", "v4", credentials=credentials, cache_discovery=False)
    return drive_service, sheets_service


def _detect_file_type(drive_service: Any, file_id: str) -> str:
    """Return the MIME type of a file in Google Drive."""
    file_meta = drive_service.files().get(fileId=file_id, fields="mimeType").execute()
    return file_meta["mimeType"]


def _download_xlsx(drive_service: Any, file_id: str) -> io.BytesIO:
    """Download an XLSX file from Google Drive into memory."""
    request = drive_service.files().get_media(fileId=file_id)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buffer.seek(0)
    return buffer


def _extract_sheet_data_xlsx(
    drive_service: Any, file_id: str, names_conversion: bool
) -> list[dict]:
    """Download XLSX and extract data from all sheets.

    Returns list of dicts: {name, headers, rows, raw_headers}
    """
    buffer = _download_xlsx(drive_service, file_id)
    wb = openpyxl.load_workbook(buffer, read_only=True, data_only=True)
    sheets_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            _log("INFO", f"Skipping empty sheet: {sheet_name}")
            continue

        raw_headers = [str(h) if h is not None else "" for h in all_rows[0]]
        data_rows = [list(row) for row in all_rows[1:]]

        raw_headers, data_rows = strip_trailing_none_columns(raw_headers, data_rows)
        if not raw_headers:
            _log("INFO", f"Skipping sheet with no valid headers: {sheet_name}")
            continue

        data_rows = filter_empty_rows(data_rows)

        if names_conversion:
            headers = deduplicate_headers([sanitize_column_name(h) for h in raw_headers])
        else:
            headers = deduplicate_headers(raw_headers)

        sheets_data.append({
            "name": sheet_name,
            "headers": headers,
            "rows": data_rows,
            "raw_headers": raw_headers,
        })

    wb.close()
    return sheets_data


def _extract_sheet_data_native(
    sheets_service: Any, spreadsheet_id: str, names_conversion: bool
) -> list[dict]:
    """Extract data from a native Google Sheets spreadsheet.

    Returns list of dicts: {name, headers, rows, raw_headers}
    """
    meta = sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheet_titles = [s["properties"]["title"] for s in meta["sheets"]]
    sheets_data = []

    for title in sheet_titles:
        result = (
            sheets_service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=title)
            .execute()
        )
        values = result.get("values", [])
        if not values:
            _log("INFO", f"Skipping empty sheet: {title}")
            continue

        raw_headers = [str(h) if h else "" for h in values[0]]
        data_rows = [list(row) for row in values[1:]]

        for i, row in enumerate(data_rows):
            if len(row) < len(raw_headers):
                data_rows[i] = row + [None] * (len(raw_headers) - len(row))

        raw_headers, data_rows = strip_trailing_none_columns(raw_headers, data_rows)
        if not raw_headers:
            continue

        data_rows = filter_empty_rows(data_rows)

        if names_conversion:
            headers = deduplicate_headers([sanitize_column_name(h) for h in raw_headers])
        else:
            headers = deduplicate_headers(raw_headers)

        sheets_data.append({
            "name": title,
            "headers": headers,
            "rows": data_rows,
            "raw_headers": raw_headers,
        })

    return sheets_data


# ---------------------------------------------------------------------------
# Stream class (for test compatibility and structure)
# ---------------------------------------------------------------------------

class GoogleSheetsStream:
    """A single sheet/tab as an Airbyte stream."""

    def __init__(self, sheet_name: str, headers: list[str], rows: list[list[Any]]):
        self._sheet_name = sheet_name
        self._headers = headers
        self._rows = rows
        self._schema = build_schema_from_sample(headers, rows)

    @property
    def name(self) -> str:
        return self._sheet_name

    def get_json_schema(self) -> dict:
        return self._schema

    def read_records(self):
        for row in self._rows:
            record = {}
            for i, header in enumerate(self._headers):
                val = row[i] if i < len(row) else None
                record[header] = serialize_value(val)
            yield record


# ---------------------------------------------------------------------------
# Airbyte Source — implements spec / check / discover / read
# ---------------------------------------------------------------------------

class SourceGoogleSheetsXlsx:
    """Airbyte source that reads Google Sheets and XLSX files from Google Drive."""

    def spec(self) -> dict:
        spec_path = Path(__file__).parent / "spec.yaml"
        with open(spec_path, "r") as f:
            return yaml.safe_load(f)

    def check_connection(self, config: dict) -> tuple[bool, Optional[str]]:
        try:
            credentials_json = config["credentials_json"]
            spreadsheet_id = config["spreadsheet_id"]

            drive_service, sheets_service = _build_google_services(credentials_json)
            mime_type = _detect_file_type(drive_service, spreadsheet_id)

            if mime_type == MIME_XLSX:
                buffer = _download_xlsx(drive_service, spreadsheet_id)
                wb = openpyxl.load_workbook(buffer, read_only=True, data_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                _log("INFO", f"XLSX file detected with {sheet_count} sheet(s)")
            elif mime_type == MIME_GOOGLE_SHEETS:
                meta = sheets_service.spreadsheets().get(
                    spreadsheetId=spreadsheet_id
                ).execute()
                sheet_count = len(meta.get("sheets", []))
                _log("INFO", f"Native Google Sheets detected with {sheet_count} sheet(s)")
            else:
                return False, f"Unsupported MIME type: {mime_type}"

            return True, None

        except Exception as e:
            return False, str(e)

    def discover(self, config: dict) -> dict:
        """Return an Airbyte catalog with all available streams."""
        streams = self._build_streams(config)
        catalog_streams = []
        for stream in streams:
            catalog_streams.append({
                "name": stream.name,
                "json_schema": stream.get_json_schema(),
                "supported_sync_modes": ["full_refresh"],
            })
        return {"streams": catalog_streams}

    def read(self, config: dict, catalog: dict, state: dict | None = None) -> None:
        """Read records and emit them as Airbyte messages."""
        configured_streams = {
            s["stream"]["name"] for s in catalog.get("streams", [])
        }
        streams = self._build_streams(config)
        emitted_at = int(time_mod.time() * 1000)

        for stream in streams:
            if stream.name not in configured_streams:
                continue
            _log("INFO", f"Reading stream: {stream.name}")
            count = 0
            for record in stream.read_records():
                _emit_record(stream.name, record, emitted_at)
                count += 1
            _log("INFO", f"Stream '{stream.name}': {count} records emitted")

        _emit_state(state or {})

    def _build_streams(self, config: dict) -> list[GoogleSheetsStream]:
        credentials_json = config["credentials_json"]
        spreadsheet_id = config["spreadsheet_id"]
        names_conversion = config.get("names_conversion", True)

        drive_service, sheets_service = _build_google_services(credentials_json)
        mime_type = _detect_file_type(drive_service, spreadsheet_id)

        if mime_type == MIME_XLSX:
            _log("INFO", "Routing to XLSX download path")
            sheets_data = _extract_sheet_data_xlsx(
                drive_service, spreadsheet_id, names_conversion
            )
        elif mime_type == MIME_GOOGLE_SHEETS:
            _log("INFO", "Routing to native Google Sheets API path")
            sheets_data = _extract_sheet_data_native(
                sheets_service, spreadsheet_id, names_conversion
            )
        else:
            raise ValueError(f"Unsupported MIME type: {mime_type}")

        streams = []
        for sheet in sheets_data:
            stream_name = (
                sanitize_column_name(sheet["name"]) if names_conversion else sheet["name"]
            )
            s = GoogleSheetsStream(
                sheet_name=stream_name,
                headers=sheet["headers"],
                rows=sheet["rows"],
            )
            streams.append(s)
            _log(
                "INFO",
                f"Stream '{stream_name}': {len(sheet['headers'])} columns, "
                f"{len(sheet['rows'])} rows",
            )

        return streams


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def run_cli(args: list[str] | None = None) -> None:
    """Parse Airbyte CLI commands and dispatch."""
    if args is None:
        args = sys.argv[1:]

    source = SourceGoogleSheetsXlsx()

    if not args:
        print("Usage: main.py {spec|check|discover|read} [--config FILE] [--catalog FILE]", file=sys.stderr)
        sys.exit(1)

    command = args[0]

    if command == "spec":
        spec = source.spec()
        _emit({"type": "SPEC", "spec": spec})
        return

    # Parse --config and --catalog
    config_path = None
    catalog_path = None
    i = 1
    while i < len(args):
        if args[i] == "--config" and i + 1 < len(args):
            config_path = args[i + 1]
            i += 2
        elif args[i] == "--catalog" and i + 1 < len(args):
            catalog_path = args[i + 1]
            i += 2
        else:
            i += 1

    if config_path is None:
        print("Error: --config is required", file=sys.stderr)
        sys.exit(1)

    with open(config_path, "r") as f:
        config = json.load(f)

    if command == "check":
        ok, error = source.check_connection(config)
        _emit({
            "type": "CONNECTION_STATUS",
            "connectionStatus": {
                "status": "SUCCEEDED" if ok else "FAILED",
                "message": error or "",
            },
        })

    elif command == "discover":
        catalog = source.discover(config)
        _emit({"type": "CATALOG", "catalog": catalog})

    elif command == "read":
        if catalog_path is None:
            print("Error: --catalog is required for read", file=sys.stderr)
            sys.exit(1)
        with open(catalog_path, "r") as f:
            catalog = json.load(f)
        source.read(config, catalog)

    else:
        print(f"Unknown command: {command}", file=sys.stderr)
        sys.exit(1)
